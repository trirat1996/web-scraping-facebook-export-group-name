
#import modules
appname = "Face plus*"
from tkinter import *

from tkinter import ttk 
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import time
import mysql.connector
import openpyxl 
from openpyxl import load_workbook


path = "configf.xlsx"

partfile_key = "key-program.txt"
part_excel_group_id = ""


def var_app():
    global read_xlsx1,r_x2,r_x3,r_x4,r_x5,r_x6,r_x7,r_x8,r_x9,r_x10 
    try:
        xlsx = openpyxl.load_workbook(path)
        xlsx_active = xlsx.active
        read_xlsx1 = xlsx_active.cell(row = 1, column = 2).value
        r_x2 = xlsx_active.cell(row = 2, column = 2).value
        r_x3 = xlsx_active.cell(row = 3, column = 2).value
        print(read_xlsx1,r_x2,r_x3)
    except:
        user_not_found("Error","พบข้อผิดพลาดบางอย่าง"," Error : configf.xlsx")




def user_not_found(w_name_log,error_log,error_code):
    global user_not_found_screen
    user_not_found_screen = Tk()
    user_not_found_screen.title(w_name_log)
    user_not_found_screen.geometry("250x150")
    Label(user_not_found_screen, text="").pack()
    Label(user_not_found_screen, text=error_log ,fg="red",font=('Tahoma', 12)).pack()
    Label(user_not_found_screen, text=error_code ,fg="red",font=('Tahoma', 9)).pack()
    Label(user_not_found_screen, text="").pack()
    Button(user_not_found_screen, text="OK", command=delete_user_not_found_screen).pack(ipadx=25, ipady=8)
    user_not_found_screen.mainloop()
 
def delete_user_not_found_screen():
    user_not_found_screen.destroy()
    
 
def facebook_login(f_menu):
    global email
    global password
    global driver
    
    emailint  = email_groupid.get()
    passwordint = pass_groupid.get()
    if f_menu == '1':
        email = r_x2
        password =  r_x3
    elif f_menu == '2':
         email = emailint 
         password =  passwordint 

    else:
        print(f_menu)

    try:
        
        driver = webdriver.Chrome() 
        driver.get('https://mbasic.facebook.com/')
        email_input = driver.find_element(By.XPATH, '//*[@id="m_login_email"]')
        time.sleep(2)
        email_input.send_keys(email)
        password_input = driver.find_element(By.XPATH, '//*[@id="password_input_with_placeholder"]/input')
        time.sleep(2)
        password_input.send_keys(password)
        submit = driver.find_element(By.XPATH, '//*[@id="login_form"]/ul/li[3]/input')  
        time.sleep(3)
        submit.click()
        time.sleep(3)
        driver.get('https://mbasic.facebook.com/')
        time.sleep(3)
        driver.find_element(By.NAME, 'xc_message').click()
        
        if f_menu == '1':
            print(f_menu)
            scan_grroup_id()
        elif f_menu == '2':
            print(f_menu)
            scan_grroup_id()
            
        else:
            print(f_menu)
    except:
        driver.quit()
        user_not_found("Error","พบข้อผิดพลาดบางอย่าง"," Error : facebook_login")
    
 
    
   
def scan_grroup_id():
    driver.get('https://m.facebook.com/groups_browse/your_groups/')
    time.sleep(2)

    iff = 1
    while iff < 20:
        iff += 1
        driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
        time.sleep(1)

    wb = openpyxl.Workbook() 
    sheet = wb.active 
    sheet.cell(row = 1, column = 1).value = "Group ID"
    sheet.cell(row= 1 , column = 2).value = "Group Name"
    sheet.cell(row= 1 , column = 3).value = "Group Link"
    remail = email[0:5]
    i = 1
    b = 1
    st = 2
    while i != 0 :
        try:
            tert = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div[2]/div/div[2]/div['+str(b)+']/a').get_attribute('href')
            texttert = driver.find_element(By.XPATH,    '//*[@id="root"]/div/div/div/div[2]/div/div[2]/div['+str(b)+']/a/div/div/div/div[1]/div').text
            #print(texttert)                        
            #print(tert[30:45])
            #print(tert)
            roop = tert.split('/')
            group_id = roop[4]
            group_name = texttert
            group_link = tert
            sheet.cell(row = st, column = 1).value = group_id
            sheet.cell(row= st , column = 2).value = group_name
            sheet.cell(row= st , column = 3).value = group_link
            #print(b)
            i = 1
            b += 1
            st += 1
        except NoSuchElementException:
            i = 0
            print("Element not found")
    part_excel_group_id_save  = str(part_excel_group_id)+"group-id-"+remail+".xlsx"      
    wb.save(part_excel_group_id_save) 
    driver.quit()
    user_not_found("Success","ดึงข้อมูลสำเร็จ",part_excel_group_id_save)       
 
 
def login_verify():
    username1 = username_verify.get()
    password1 = password_verify.get()
    username_login_entry.delete(0, END)
    password_login_entry.delete(0, END)
 
    list_of_files = os.listdir()
    if username1 in list_of_files:
        file1 = open(username1, "r")
        verify = file1.read().splitlines()
        if password1 in verify:
            login_sucess()
 
        else:
            login_sucess()
 
    else:
        login_sucess()
 
# Designing popup for login success
 
def login_sucess():
    global login_success_screen
    login_success_screen = Toplevel(login_screen)
    login_success_screen.title("Success")
    login_success_screen.geometry("150x100")
    Label(login_success_screen, text="Login Success").pack()
    Button(login_success_screen, text="OK", command=delete_login_success).pack()
 
# Designing popup for login invalid password
 

 
def delete_login_success():
    login_success_screen.destroy()
    
def save_main_user():
    var_app()
    main_user1 = email_groupid.get()
    main_pass1 = pass_groupid.get()
    if  main_user1 =='':
        main_user = ' null '
        main_pass = ' null  '
    else :
        main_user = main_user1.replace(' ', '')
        main_pass = main_pass1.replace(' ', '')
    up_xlsx = load_workbook(path)

    up_xlsx2 = up_xlsx.active
    up_xlsx2["B2"] = main_user
    up_xlsx2["B3"] = main_pass
    up_xlsx.save(path)
    root.destroy()
    chacklisen()
 


def login():
   #main_screen = Tk()
    global login_screen
    #login_screen = Toplevel(main_screen)
    login_screen =  Tk()
    login_screen.title("Login")
    login_screen.geometry("300x250")
    Label(login_screen, text="Please enter details below to login").pack()
    Label(login_screen, text="").pack()
 
    global username_verify
    global password_verify
 
    username_verify = StringVar()
    password_verify = StringVar()
 
    global username_login_entry
    global password_login_entry
 
    Label(login_screen, text="Username * ").pack()
    username_login_entry = Entry(login_screen, textvariable=username_verify)
    username_login_entry.pack()
    Label(login_screen, text="").pack()
    Label(login_screen, text="Password * ").pack()
    password_login_entry = Entry(login_screen, textvariable=password_verify, show= '*')
    password_login_entry.pack()
    Label(login_screen, text="").pack()
    Button(login_screen, text="Login", width=10, height=1, command = main_menu).pack()
    Label(login_screen, text="").pack()
    Button(login_screen, text="start", width=10, height=1, command = facebook_login).pack()
    
    login_screen.mainloop()
    
def main_menu():

    try:   
        global root
        root = Tk()
    
        root.title(appname) 
        root.geometry("550x500")
        tabControl = ttk.Notebook(root) 
        
        tab1 = ttk.Frame(tabControl) 
        tab2 = ttk.Frame(tabControl) 
        
        
        tabControl.add(tab1, text ="Group ID") 
        tabControl.add(tab2, text ='Tab 2') 
        tabControl.pack(expand = 1, fill ="both") 
        
        global email_groupid
        global pass_groupid
        
        email_groupid = StringVar()
        pass_groupid = StringVar()

        Label(tab1,text=r"ดึงข้อมูล ชือ,ID กลุ่มทั้งหมดที่เป็นสมาชิก",fg="blue",font=('Tahoma', 16)).grid(row = 0, column = 1, columnspan = 4,pady = 15)
        textpartxls = "วิธีใช้งาน : ใส่ข้อมูล E-mail และ Password ของเฟสบุ๊ค โปรแกรมจะดึงข้อมูลแล้ว ไปเก็บไว้ที่ "+part_excel_group_id
        ttk.Label(tab1,text=textpartxls ,font=('Tahoma', 10)).grid(row = 1, column = 1,columnspan = 4)
        ttk.Label(tab1,text="หากต้องการบันทึกข้อมูล  E-mail และ Passwordไว้ใช้ครั้งถัดไปกด บันทึก",font=('Tahoma', 10)).grid(row = 3, column = 1, columnspan = 4)
        ttk.Label(tab1,text="และหากต้องการเรียกใช้ข้อมูลล่าสุด กด เริ่มจากบันทึก",font=('Tahoma', 10)).grid(row = 4, column = 1,columnspan = 4)
        if r_x2 =='': r_x2x = "null" 
        else: r_x2x = r_x2  
        namesave = "ชื่อที่บันทึกไว้ล่าสุด : " + r_x2x
        Label(tab1,text=namesave,fg="gray",font=('Tahoma', 10)).grid(row = 5, column = 1, columnspan = 4,pady = 20)
        
        ttk.Label(tab1, text="  ").grid(row = 6, column = 0, sticky = W, pady = 10 ,padx=5)
        ttk.Label(tab1, text="E-mail * ").grid(row = 6, column = 1, sticky = W, pady = 5 ,padx=5)
        username_login_entry = Entry(tab1, textvariable = email_groupid)
        
        ttk.Label(tab1, text="Password * ").grid(row = 7, column = 1, sticky = W, pady = 5 ,padx=5)
        password_login_entry = Entry(tab1,  textvariable = pass_groupid)
        username_login_entry.grid(row = 6, column = 2, sticky = W,pady = 5 ,padx=5,ipadx=50, ipady=2)
        password_login_entry.grid(row = 7, column = 2, sticky = W,pady = 5 ,padx=5,ipadx=50, ipady=2)
    # Label(tab1,text=r"____________________________________________________________________________________________________________________________________",fg="gray").grid(sticky = W,row = 8, column = 0,columnspan = 6)
        Button(tab1, text=r"บันทึก", width=10, height=1, fg="green",font=('Tahoma', 8), command = save_main_user).grid(row = 6, column = 3, pady = 5 ,padx=5)
        Button(tab1, text="เริ่มจากบันทึก", width=10, height=1, font=('Tahoma', 8), command = lambda: facebook_login('1')).grid(row = 7, column = 3, pady = 5 ,padx=5)
        Button(tab1, text="เริ่มทำงาน", width=10, height=1, font=('Tahoma', 8), command =lambda:  facebook_login('2')).grid(row = 9, column = 3, pady = 5 ,padx=5)

        ttk.Label(tab2, text ="Lets dive into the\ world of computers").pack()
        root.mainloop()  
    except:user_not_found("Error","พบข้อผิดพลาดบางอย่าง"," Error : main_menu")
    
def enterkey():

    addkertmsg = addkey.get()
    addkert = addkertmsg.replace(' ', '')
    
    up_xlsx = load_workbook(path)

    up_xlsx1 = up_xlsx.active
    up_xlsx1["B1"] = addkert
    up_xlsx.save(path)
    
    
    #fkey = open(partfile_key, "w")
    #fkey.write(addkert)
    #fkey.close()
    keychack_screen.destroy()
    chacklisen()
    
 

def keychack():

    global keychack_screen
    
    keychack_screen =  Tk()
    keychack_screen.title(appname)
    keychack_screen.geometry("300x250")
    Label(keychack_screen, text="").pack()
    Label(keychack_screen, text="กรุณาใส่รหัสใช้งานโปรแกรม",font=('Tahoma', 10, 'bold')).pack()
    Label(keychack_screen, text="").pack()
 
    global addkey
    addkey = StringVar()
    
    global addkey_entry
    Label(keychack_screen, text="KEY").pack()
    addkey_entry = Entry(keychack_screen, textvariable = addkey)
    addkey_entry.pack()
    Label(keychack_screen, text="").pack()
    Button(keychack_screen, text="Inspect", width=10, height=1, command = enterkey).pack()
    Label(keychack_screen, text="").pack()
    Label(keychack_screen, text="ติดต่อเพื่อใช้งานโปรแกรมได้ที่",font=('Tahoma', 8)).pack()
    Label(keychack_screen, text=r"http://www.google.com", fg="blue", cursor="hand2").pack()
    keychack_screen.mainloop()
    

    
    
    
def datakey():
    var_app()
    try:
        
        keyprogram = "('"+read_xlsx1+"',)"
        #with open(partfile_key , encoding="utf8") as fkey:
        #keyp = fkey.read() 
        #keyprogram = "('"+keyp+"',)"
    except:
        user_not_found("Error","พบข้อผิดพลาดบางอย่าง"," Error : key_chack.txt")
        keychack() 
    #print(keyprogram)
    
    
    try:
        dataBase = mysql.connector.connect(
        host ="sql10.freesqldatabase.com",
        user ="sql10666189",
        passwd ="MIm9V4qVHr",
        database = "sql10666189")
# preparing a cursor object
        cursorObject = dataBase.cursor()
        query = "SELECT userkey FROM keychack WHERE active = '1' "
        cursorObject.execute(query)
        ckey = cursorObject.fetchall()
        print(ckey)
# disconnecting from server
        dataBase.close()
    except:
        user_not_found("Error","พบข้อผิดพลาดบางอย่าง"," Error : Sql Not Nonnect")   
    global keysu
    keysu = str(keyprogram) in str(ckey)
    print(keysu)    

def chacklisen():
    datakey()
    try:
        if  keysu == True :  main_menu()
        else:keychack()
    except:
        #if  keysu != True : not_connet()
        user_not_found("Error","พบข้อผิดพลาดบางอย่าง"," ***")   
        #keychack()    

    


#datakey()
chacklisen()


